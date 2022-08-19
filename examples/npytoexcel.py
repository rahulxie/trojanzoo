# 直接上代码
import numpy as np
import openpyxl
from openpyxl import Workbook
import pandas as pd
import torch

title = '../data/quant/quant/defense_estimation_epoch150_clean_q5/image/gtsrb/resnet18_comp/neural_cleanse/badnet_square_white_tar0_alpha1.00_mark(32,32)'  # 读取的文件名
data = np.load(title + '.npz')  # 读取numpy文件
mask_norms = torch.from_numpy(data['mark_list'][:, -1]).flatten(start_dim=1).norm(p=1, dim=1)
mask_norm_list = mask_norms.tolist()
sheet_name = 'gg'

df = pd.DataFrame(mask_norm_list, columns=['mask_norms'])
path_excel = "mask_norms_q.xlsx"
wb = openpyxl.load_workbook(path_excel)  # 将空白表格指定为wb
# wb.create_sheet(title=sheet_name, index=0)

writer = pd.ExcelWriter(path_excel, engine='openpyxl')  # 准备往空白表格写入数据
writer.book = wb  # 执行写入目标为空白表格
df.to_excel(writer, sheet_name=sheet_name)
writer.save()
writer.close()
# wb = Workbook()
# ws = wb.create_sheet("cifar10")
